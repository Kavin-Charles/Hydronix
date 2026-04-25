"""
Build the Hydronix hackathon-submission bundle.

Produces under ./submission/:
    Hydronix_Code.zip          full source tree (hydro/ + main.py + app.py + samples)
    Hydronix_Demo.zip          same payload, plus run scripts (executable demo)
    Hydronix_Report.pdf        PDF report (PS hull, T = 28.5 m, KG = 18 m)
    Hydronix_Output.xlsx       computed hydrostatic values + GZ / KN tables
    Hydronix_GZ_Curve.png      true-heeled GZ vs heel angle
    Hydronix_KN_Curve.png      KN cross-curves at multiple displacements

Run:
    python scripts/build_submission.py
"""

from __future__ import annotations

import os
import shutil
import sys
import zipfile
from pathlib import Path

os.environ.setdefault("MPLBACKEND", "Agg")

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from hydro.benchmarks    import hackathon_ps_hull
from hydro.hydrostatics  import Hydrostatics, hydrostatic_table
from hydro.heeled        import gz_curve_true, kn_curve_true, cross_curves_true
from hydro.stability     import gz_curve_wallsided, stability_parameters
from hydro.imo           import imo_intact_stability_check
from hydro              import plots
from hydro.report        import generate_pdf


SUB = ROOT / "submission"
FIG = SUB / "figures"


def _mkdirs() -> None:
    for d in (SUB, FIG):
        d.mkdir(exist_ok=True)


def _gen_artifacts():
    print("[1/5] Loading PS hull and computing hydrostatics …")
    hull   = hackathon_ps_hull()
    draft  = 28.5
    KG     = 18.0
    rho    = 1.025

    hs     = Hydrostatics(hull, draft=draft, KG=KG)
    s      = hs.summary()
    s["KG_m"] = KG

    angles = np.arange(0.0, 81.0, 5.0)
    ang, gz = gz_curve_true(hull, draft=draft, KG=KG, angles_deg=angles)
    _, kn   = kn_curve_true(hull, draft=draft, KG=KG, angles_deg=angles)
    ang_w, gz_w = gz_curve_wallsided(hs, angles_deg=angles)
    params  = stability_parameters(ang, gz, hs.GM)
    imo     = imo_intact_stability_check(params)

    print("[2/5] Plotting GZ + KN curves …")
    gz_path = SUB / "Hydronix_GZ_Curve.png"
    fig_gz = plots.plot_gz_curve(ang, gz, ang_w, gz_w, params,
                                 ship_name=hull.name)
    fig_gz.savefig(gz_path, dpi=200, bbox_inches="tight")
    plt.close(fig_gz)

    # KN cross-curves at five draughts spanning 60–130 % of design draft
    cc_drafts = np.array([0.6, 0.8, 1.0, 1.15, 1.3]) * draft
    cc_drafts = [float(t) for t in cc_drafts if t < hull.D * 0.95]
    cc = cross_curves_true(hull, cc_drafts, KG=KG,
                           angles_deg=[10, 20, 30, 40, 50, 60, 70, 80])
    # kn_matrix from cross_curves_true is shape (n_angles, n_drafts).
    # plot_kn_curves expects [n_drafts][n_angles] → transpose.
    kn_per_draft = np.array(cc["kn_matrix"]).T.tolist()
    kn_path = SUB / "Hydronix_KN_Curve.png"
    fig_kn = plots.plot_kn_curves(np.array(cc["angles_deg"]),
                                  kn_per_draft,
                                  cc["displacements_t"],
                                  ship_name=hull.name)
    fig_kn.savefig(kn_path, dpi=200, bbox_inches="tight")
    plt.close(fig_kn)

    print("[3/5] Writing computed-values workbook …")
    xlsx_path = SUB / "Hydronix_Output.xlsx"

    # Metadata / cover sheet
    meta_rows = [
        ("Project",            "Hydronix — Ship Hydrostatics Suite"),
        ("Authors",            "Kavin Charles · Jeevika R"),
        ("Event",              "Wavez 2026 · IIT Madras"),
        ("Hull",               hull.name),
        ("LOA / LBP (m)",      f"{hull.L:.2f}"),
        ("Maximum breadth B (m)", f"{hull.B_max:.2f}"),
        ("Moulded depth D (m)",  f"{hull.D:.2f}"),
        ("Design draft T (m)", f"{draft:.2f}"),
        ("KG (m)",             f"{KG:.2f}"),
        ("Density rho (t/m^3)",f"{rho:.3f}"),
        ("IMO 2008 IS verdict", imo["overall"]),
        ("Criteria passed",    f"{imo['passed']} / {imo['passed']+imo['failed']}"),
    ]
    meta_df = pd.DataFrame(meta_rows, columns=["Field", "Value"])

    # Sub-sheets — friendly column names + units
    pretty_keys = {
        "displacement_m3":    "Displacement volume Vol (m^3)",
        "displacement_t":     "Displacement Disp (t)",
        "waterplane_area_m2": "Waterplane area Aw (m^2)",
        "Am_m2":              "Midship section area Am (m^2)",
        "TPC_t_per_cm":       "TPC (t/cm)",
        "MCTC_tm_per_cm":     "MCTC (t.m/cm)",
        "lcb_from_ap_m":      "LCB from AP (m)",
        "lcf_from_ap_m":      "LCF from AP (m)",
        "KB_m":               "KB (m)",
        "BM_m":               "BM (m)",
        "KM_m":               "KM (m)",
        "KG_m":               "KG (m)",
        "GM_m":               "GM (m)",
        "BML_m":              "BML (m)",
        "GML_m":              "GML (m)",
        "IT_m4":              "IT — transverse (m^4)",
        "IL_m4":              "IL — longitudinal (m^4)",
        "Cb":                 "Block coefficient Cb",
        "Cw":                 "Waterplane coefficient Cw",
        "Cm":                 "Midship coefficient Cm",
        "Cp":                 "Prismatic coefficient Cp",
        "Cvp":                "Vertical prismatic Cvp",
        "roll_period_s":      "Natural roll period Tphi (s)",
        "L_m":                "Length L (m)",
        "B_max_m":             "Max breadth B (m)",
        "D_m":                "Depth D (m)",
        "draft_m":             "Design draft T (m)",
        "rho_t_per_m3":        "Density rho (t/m^3)",
    }
    summary_rows = []
    for k, v in s.items():
        label = pretty_keys.get(k, k)
        if isinstance(v, float):
            v = round(v, 6)
        summary_rows.append((label, v))
    summary_df = pd.DataFrame(summary_rows, columns=["Quantity", "Value"])

    gz_df = pd.DataFrame({
        "phi_deg":          np.round(ang, 3),
        "GZ_true_m":        np.round(gz, 6),
        "GZ_wallsided_m":   np.round(gz_w, 6),
        "KN_m":             np.round(kn, 6),
    })
    T_grid = np.linspace(0.2 * hull.D, 0.95 * hull.D, 12)
    table  = hydrostatic_table(hull, T_grid.tolist(), KG=KG)
    table_df = pd.DataFrame(table).round(4)
    imo_df = pd.DataFrame(imo["criteria"])

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as xl:
        meta_df    .to_excel(xl, sheet_name="00 — Cover",         index=False)
        summary_df .to_excel(xl, sheet_name="01 — Hydrostatics",  index=False)
        table_df   .to_excel(xl, sheet_name="02 — Hydrostatic Table", index=False)
        gz_df      .to_excel(xl, sheet_name="03 — GZ & KN Curves", index=False)
        imo_df     .to_excel(xl, sheet_name="04 — IMO 2008 IS",    index=False)

    # Polish formatting: header style, column widths, freeze panes
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = load_workbook(xlsx_path)
    header_fill = PatternFill("solid", fgColor="1F6AA5")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    body_font   = Font(size=10)
    centre      = Alignment(horizontal="center", vertical="center")
    left_align  = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
            for cell in col:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = centre
        # Auto-fit-ish: widen columns to 1.2x longest cell length
        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) for c in col_cells if c.value is not None),
                          default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(48, max(12, max_len + 2))
        # Body cells light styling
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = centre
                    if isinstance(cell.value, float):
                        cell.number_format = "#,##0.0000"
                else:
                    cell.alignment = left_align
        ws.row_dimensions[1].height = 22

    # Cover sheet wider on first column
    cover = wb["00 — Cover"]
    cover.column_dimensions["A"].width = 30
    cover.column_dimensions["B"].width = 50
    wb.save(xlsx_path)

    print("[4/5] Generating PDF report (with all figures) …")
    body_path     = FIG / "body_plan.png"
    curves_path   = FIG / "curves_of_form.png"
    bonjean_path  = FIG / "bonjean.png"
    imo_fig_path  = FIG / "imo.png"
    plots.plot_body_plan(hull, draft, ship_name=hull.name,
                         save_path=str(body_path))
    plots.plot_hydrostatic_curves(table, ship_name=hull.name,
                                  save_path=str(curves_path))
    from hydro.bonjean import bonjean_curves
    bj = bonjean_curves(hull, dT=max(0.1, hull.D / 30))
    plots.plot_bonjean(bj, ship_name=hull.name,
                       save_path=str(bonjean_path))
    plots.plot_imo_criteria(imo, ship_name=hull.name,
                            save_path=str(imo_fig_path))

    pdf_path = SUB / "Hydronix_Report.pdf"
    generate_pdf(
        output_path   = pdf_path,
        ship_name     = hull.name,
        hydro_summary = s,
        stab_params   = params,
        imo_check     = imo,
        figure_paths  = {
            "body_plan": str(body_path),
            "gz"       : str(gz_path),
            "curves"   : str(curves_path),
            "bonjean"  : str(bonjean_path),
            "imo"      : str(imo_fig_path),
        },
    )

    print("[5/5] Bundling source code zips …")
    code_zip = SUB / "Hydronix_Code.zip"
    demo_zip = SUB / "Hydronix_Demo.zip"

    code_paths = [
        ROOT / "hydro",
        ROOT / "samples",
        ROOT / "main.py",
        ROOT / "app.py",
        ROOT / "requirements.txt",
        ROOT / "README.md",
        ROOT / "tests",
        ROOT / "assets",
    ]
    demo_paths = code_paths + [ROOT / "scripts"]

    def _zip(out: Path, paths: list[Path]) -> None:
        if out.exists():
            out.unlink()
        with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
            for p in paths:
                if not p.exists():
                    continue
                if p.is_file():
                    zf.write(p, arcname=p.relative_to(ROOT))
                else:
                    for f in p.rglob("*"):
                        if f.is_file() and "__pycache__" not in f.parts:
                            zf.write(f, arcname=f.relative_to(ROOT))

    _zip(code_zip, code_paths)
    _zip(demo_zip, demo_paths)

    print()
    print("=" * 60)
    print("Submission bundle ready in submission/")
    for f in sorted(SUB.iterdir()):
        if f.is_file():
            kb = f.stat().st_size / 1024
            print(f"  {f.name:40s}  {kb:9.1f} KB")
    print("=" * 60)


if __name__ == "__main__":
    _mkdirs()
    _gen_artifacts()
