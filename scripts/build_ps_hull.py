"""
Convert the official Wavez 2026 Hydrohackathon Problem-Statement offset
table (`69ec3fe06117b_HydroHackathon_PS_File.xlsx`) into the project's
native Hull JSON format and drop it under `samples/`.

Run:
    python scripts/build_ps_hull.py
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import numpy as np
import openpyxl


ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "69ec3fe06117b_HydroHackathon_PS_File.xlsx"
OUT  = ROOT / "samples" / "hackathon_ps_hull.json"


def main() -> int:
    if not XLSX.exists():
        print(f"[err] cannot find {XLSX}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Offset data"]

    # Stations: row 4, cols D..Z (4..26) — metric x positions
    stations: list[float] = []
    for c in range(4, 27):
        v = ws.cell(row=4, column=c).value
        if v is not None:
            stations.append(float(v))

    # Waterlines:
    #   row 6  → keel z=0
    #   rows 7..16 → A..K with z value in column 3
    waterlines: list[float] = []
    half_breadths_per_wl: list[list[float]] = []

    for r in range(6, 17):
        z_val = ws.cell(row=r, column=3).value
        z = 0.0 if r == 6 else float(z_val)
        waterlines.append(z)
        row_vals: list[float] = []
        for c in range(4, 27):
            v = ws.cell(row=r, column=c).value
            row_vals.append(0.0 if v is None else float(v))
        half_breadths_per_wl.append(row_vals)

    # Reshape: hb[station, waterline]
    hb = np.asarray(half_breadths_per_wl, dtype=float).T

    out = {
        "name":          "Hackathon PS Container Hull",
        "stations":      list(map(float, stations)),
        "waterlines":    list(map(float, waterlines)),
        "half_breadths": hb.tolist(),
        "rho":           1.025,
        "_meta": {
            "source": str(XLSX.name),
            "LOA_m":   420.95,
            "LBP_m":   stations[-1] - stations[0],
            "B_m":     63.0,
            "T_design_m": 28.5,
            "D_m":     37.269,
            "CB_reported": 0.78,
        },
    }
    OUT.parent.mkdir(exist_ok=True)
    OUT.write_text(json.dumps(out, indent=2))
    print(f"wrote {OUT}")
    print(f"  stations  : {len(stations)}")
    print(f"  waterlines: {len(waterlines)}")
    print(f"  hb shape  : {hb.shape}")
    print(f"  hb max    : {hb.max():.3f} m  (B/2 expected = 31.5)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
